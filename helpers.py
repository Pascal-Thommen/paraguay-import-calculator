"""
helpers.py – translations, HS‑code rules, and state persistence
for the Paraguay Import Cost Calculator.

.. deprecated::
    Die Berechnungsfunktionen ``calc_single_product`` und ``calc_multi_product``
    wurden nach :mod:`calculator` ausgelagert. Verwende ``calculator.py`` stattdessen.
"""
import streamlit as st
import json
import os
import re

# =====================================================================
# Persistent state file (alongside the app)
# =====================================================================
STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".app_state.json")

# =====================================================================
# Default (example) values – shown on first load & after reset
# =====================================================================
DEFAULTS = {
    "lang": "de",
    # Single‑product
    "p_name": "Laptop",
    "hs_code": "84713000",
    "p_qty": 100,
    "p_fob_usd": 450.0,
    "p_weight": 2.5,
    "freight_usd": 3500.0,
    "insurance_usd": 400.0,
    "inland_pyg": 3000000.0,
    "inland_iva_incl": True,
    "dai_rate": 6.0,
    "val_mode": 0,          # 0 = auto 0.5 %, 1 = manual
    "val_pyg_input": 0.0,
    "indi_rate": 7.0,
    "canon_sofia": 100000.0,
    "consulado": 500000.0,
    "tasa_portuaria": 1500000.0,
    "despachante": 2500000.0,
    "despachante_iva_incl": True,
    "sonstiges": 300000.0,
    # Sidebar globals
    "ex_rate": 7700.0,
    "percepcion_ire_rate": 0.4,
    # Multi‑product shared costs
    "multi_freight_usd": 3500.0,
    "multi_insurance_usd": 400.0,
    "multi_inland_pyg": 3000000.0,
    "multi_inland_iva_incl": True,
    "multi_val_mode": 0,
    "multi_val_pyg_manual": 800000.0,
    "multi_indi_rate": 7.0,
    "multi_canon_sofia": 100000.0,
    "multi_consulado": 500000.0,
    "multi_tasa_portuaria": 1500000.0,
    "multi_despachante": 2500000.0,
    "multi_despachante_iva_incl": True,
    "multi_sonstiges": 300000.0,
    "alloc_freight": 0,   # 0 = weight, 1 = value
    "alloc_local": 0,     # 0 = value,  1 = weight
}

# Keys that we persist / reset (scalars only – DataFrame handled separately)
STATE_KEYS = list(DEFAULTS.keys())

# =====================================================================
# TRANSLATIONS  (de / en / es)
# =====================================================================
TRANSLATIONS: dict[str, dict[str, str]] = {
    # -----------------------------------------------------------------
    "de": {
        # Sidebar
        "lang_selector": "🌐 Sprache",
        "zone_label": "Herkunftszone",
        "zone_help": "Beeinflusst DAI-Zollsatz und Frachtschaetzungen je nach Herkunftsregion",
        "sidebar_global_params": "### 🌍 Globale Parameter",
        "exchange_rate_label": "Wechselkurs (PYG/USD)",
        "exchange_rate_help": "Aktueller Wechselkurs Paraguayischer Guaraní zu US‑Dollar.",
        "percepcion_ire_label": "Percepción IRE (%)",
        "legal_notice": "📋 Dieses Tool dient als Kalkulationshilfe. Die Ergebnisse ersetzen keine professionelle Steuer‑ oder Zollberatung.",
        "reset_button": "🔄 Alle Felder zurücksetzen",
        # Banner
        "banner_title": "🇵🇾 Import‑Kostenkalkulator Paraguay",
        "banner_subtitle": "Rechtskonforme Anschaffungskostenermittlung nach IAS 2 & Ley 6380/19",
        # Tabs
        "tab_single": "📦 Einzelprodukt‑Kalkulator",
        "tab_multi": "🗃️ Mehrprodukt‑Kalkulator",
        # Single – card
        "single_card_header": "Kalkulationsdaten für ein einzelnes Produkt",
        # Col 1
        "col1_header": "**1. Produktspezifikationen**",
        "product_name_label": "Produktname",
        "hs_code_label": "HS‑Code",
        "product_qty_label": "Menge (Stück)",
        "fob_price_label": "FOB‑Preis pro Stück (USD)",
        "product_weight_label": "Gewicht pro Stück (kg)",
        "hs_code_invalid": "⚠️ HS‑Code ungültig – bitte 6‑10 Ziffern eingeben.",
        # Col 2
        "col2_header": "**2. Logistik & Transport**",
        "freight_label": "Internationale Fracht (USD)",
        "insurance_label": "Transportversicherung (USD)",
        "inland_transport_label": "Inlandtransport (PYG)",
        "inland_iva_checkbox": "Inlandstransport enthält 10 % IVA",
        # Col 3
        "col3_header": "**3. Zoll & Abfertigung (PYG)**",
        "dai_rate_label": "DAI Zollsatz (%)",
        "valoracion_label": "Valoración Aduanera Gebühr",
        "valoracion_opt_auto": "0,5 % des CIF‑Werts",
        "valoracion_opt_manual": "Manueller Betrag",
        "valoracion_manual_label": "Manueller Betrag Valoración (PYG)",
        "indi_rate_label": "INDI Satz (% von DAI)",
        "canon_sofia_label": "Canon SOFIA (PYG)",
        "consulado_label": "Konsulatsgebühren (PYG)",
        "tasa_portuaria_label": "Hafengebühren (PYG)",
        "despachante_label": "Despachante Honorar (PYG)",
        "despachante_iva_checkbox": "Despachante‑Honorar enthält 10 % IVA",
        "other_costs_label": "Sonstige Nebenkosten (PYG)",
        # Single results
        "results_header": "### 📊 Kalkulationsergebnisse (Einzelprodukt)",
        "unit_cost_title": "Stückeinstandspreis (IAS 2 aktiviert)",
        "unit_cost_sub": "Entspricht <b>{usd}</b> pro Stück (Gesamte Anschaffungskosten: {total})",
        "tax_credit_title": "Steuerliches Guthaben (Crédito Fiscal – nicht aktiviert)",
        "tax_credit_sub": "Wird direkt mit der Steuererklärung verrechnet und mindert nicht die Marge.",
        # Multi – card & intro
        "multi_card_header": "Mehrprodukt‑Importkalkulation mit Kostenschlüsselung",
        "multi_intro": "Geben Sie hier Ihre Produktliste ein. Die Nebenkosten können Sie wert‑ (FOB) oder gewichtsbasiert (kg) auf die Produkte verteilen.",
        "multi_products_header": "**1. Importierte Produkte**",
        "multi_col_product": "Produktname",
        "multi_col_hscode": "HS‑Code",
        "multi_col_qty": "Menge",
        "multi_col_fob": "FOB pro Stk. (USD)",
        "multi_col_weight": "Gewicht pro Stk. (kg)",
        "multi_col_dai": "DAI (%)",
        # Multi – logistics
        "multi_logistics_header": "**Logistikkosten (Gesamtlieferung)**",
        "multi_freight_label": "See‑/Luftfracht gesamt (USD)",
        "multi_insurance_label": "Versicherung gesamt (USD)",
        "multi_inland_label": "Inlandstransport gesamt (PYG)",
        "multi_inland_iva": "Inlandstransport enthält 10 % IVA",
        # Multi – customs
        "multi_customs_header": "**Gemeinsame Zollgebühren (PYG)**",
        "multi_valoracion_label": "Valoración Aduanera",
        "multi_val_opt_auto": "0,5 % des CIF‑Werts je Produkt",
        "multi_val_opt_manual": "Manueller Gesamtbetrag",
        "multi_val_manual_label": "Gesamt‑Valoración (PYG)",
        "multi_indi_label": "INDI Satz (% von DAI)",
        "multi_canon_label": "Canon SOFIA gesamt (PYG)",
        "multi_consulado_label": "Konsulatsgebühren gesamt (PYG)",
        "multi_tasa_label": "Hafengebühren gesamt (PYG)",
        # Multi – services
        "multi_services_header": "**Dienstleistungen & Sonstiges**",
        "multi_despachante_label": "Despachante Honorar gesamt (PYG)",
        "multi_despachante_iva": "Broker‑Honorar enthält 10 % IVA",
        "multi_sonstiges_label": "Sonstige Nebenkosten gesamt (PYG)",
        # Multi – allocation
        "multi_alloc_header": "**Verteilungsschlüssel**",
        "multi_alloc_freight_label": "Verteilung Fracht & Versicherung",
        "multi_alloc_freight_opt_weight": "Gewichtsbasiert (kg‑Anteil)",
        "multi_alloc_freight_opt_value": "Wertbasiert (FOB‑Anteil)",
        "multi_alloc_freight_help": "Frachtkosten fallen meist gewichts‑ oder volumenbezogen an.",
        "multi_alloc_local_label": "Verteilung Hafen, Broker, Inland",
        "multi_alloc_local_opt_value": "Wertbasiert (FOB‑Anteil)",
        "multi_alloc_local_opt_weight": "Gewichtsbasiert (kg‑Anteil)",
        "multi_alloc_local_help": "So werden die fixen lokalen Gebühren verteilt.",
        # Multi – warnings
        "multi_warn_no_products": "⚠️ Bitte fügen Sie mindestens ein Produkt hinzu.",
        "multi_error_zero": "Gesamt‑FOB‑Wert und Gesamtgewicht müssen > 0 sein.",
        # Multi – results
        "multi_results_header": "### 📊 Gesamtergebnisse (Mehrprodukt‑Import)",
        "multi_cap_title": "Summe Anschaffungskosten (IAS 2 aktiviert)",
        "multi_cap_sub": "Entspricht ca. <b>{usd}</b>",
        "multi_credit_title": "Summe Steuerliches Guthaben (nicht aktiviert)",
        "multi_credit_sub": "Zoll‑IVA ({iva}) + Percepción ({perc}) + Service‑IVA ({svc})",
        "multi_table_header": "**Ergebnis‑Übersicht je Produkt**",
        "multi_csv_button": "📥 Verteilungsbogen als CSV exportieren",
        "multi_xlsx_button": "📥 Verteilungsbogen als Excel exportieren",
        "multi_csv_filename": "import_kalkulation_mehrprodukt.csv",
        "multi_xlsx_filename": "import_kalkulation_mehrprodukt.xlsx",
        "multi_chart_products": "**Anschaffungskosten nach Produkt**",
        "multi_chart_types": "**Verteilung nach Kostenart**",
        "multi_legend_products": "Produkte",
        "multi_legend_types": "Kostengruppen",
        "multi_chart_fob": "FOB Warenwert",
        "multi_chart_freight": "Int. Fracht & Vers.",
        "multi_chart_customs": "Zoll & Abgaben",
        "multi_chart_local": "Lokale Logistik/Broker",
        # Result table columns
        "res_col_product": "Produkt",
        "res_col_qty": "Menge",
        "res_col_fob": "FOB USD (Ges.)",
        "res_col_weight": "Gewicht (Ges.)",
        "res_col_cif": "CIF (PYG)",
        "res_col_dai": "Zoll DAI (PYG)",
        "res_col_other": "Nebenk. (PYG)",
        "res_col_total": "AK gesamt (PYG)",
        "res_col_unit_pyg": "Stückkosten (PYG)",
        "res_col_unit_usd": "Stückkosten (USD)",
        "res_col_credit": "Steuerguthaben (PYG)",
        # Footer
        "footer_header": "### 📚 Begriffserklärungen & Steuerliche Grundlagen (Paraguay)",
    },
    # -----------------------------------------------------------------
    "en": {
        "lang_selector": "🌐 Language",
        "zone_label": "Origin Zone",
        "zone_help": "Affects DAI tariff rate and freight estimates by origin region",
        "sidebar_global_params": "### 🌍 Global Parameters",
        "exchange_rate_label": "Exchange rate (PYG/USD)",
        "exchange_rate_help": "Current exchange rate Paraguayan Guaraní to US Dollar.",
        "percepcion_ire_label": "Percepción IRE (%)",
        "legal_notice": "📋 This tool is a calculation aid. Results do not replace professional tax or customs advice.",
        "reset_button": "🔄 Reset all fields",
        "banner_title": "🇵🇾 Paraguay Import Cost Calculator",
        "banner_subtitle": "IAS 2 & Ley 6380/19 compliant acquisition‑cost calculation",
        "tab_single": "📦 Single‑product calculator",
        "tab_multi": "🗃️ Multi‑product calculator",
        "single_card_header": "Calculation data for a single product",
        "col1_header": "**1. Product specifications**",
        "product_name_label": "Product name",
        "hs_code_label": "HS code",
        "product_qty_label": "Quantity (units)",
        "fob_price_label": "FOB price per unit (USD)",
        "product_weight_label": "Weight per unit (kg)",
        "hs_code_invalid": "⚠️ Invalid HS code – please enter 6‑10 digits.",
        "col2_header": "**2. Logistics & transport**",
        "freight_label": "International freight (USD)",
        "insurance_label": "Transport insurance (USD)",
        "inland_transport_label": "Inland transport (PYG)",
        "inland_iva_checkbox": "Inland transport includes 10 % IVA",
        "col3_header": "**3. Customs & clearance (PYG)**",
        "dai_rate_label": "DAI tariff (%)",
        "valoracion_label": "Customs valuation fee",
        "valoracion_opt_auto": "0.5 % of CIF value",
        "valoracion_opt_manual": "Manual amount",
        "valoracion_manual_label": "Manual valuation amount (PYG)",
        "indi_rate_label": "INDI rate (% of DAI)",
        "canon_sofia_label": "Canon SOFIA (PYG)",
        "consulado_label": "Consular fees (PYG)",
        "tasa_portuaria_label": "Port fee (PYG)",
        "despachante_label": "Customs broker fee (PYG)",
        "despachante_iva_checkbox": "Broker fee includes 10 % IVA",
        "other_costs_label": "Other overhead (PYG)",
        "results_header": "### 📊 Calculation results (single product)",
        "unit_cost_title": "Unit acquisition cost (IAS 2 capitalised)",
        "unit_cost_sub": "Equals <b>{usd}</b> per unit (total acquisition cost: {total})",
        "tax_credit_title": "Tax credit (Crédito Fiscal – not capitalised)",
        "tax_credit_sub": "Offset directly against the tax return; does not reduce margin.",
        "multi_card_header": "Multi‑product import calculation with cost allocation",
        "multi_intro": "Enter your product list. You can allocate shared costs by value (FOB) or weight (kg).",
        "multi_products_header": "**1. Imported products**",
        "multi_col_product": "Product",
        "multi_col_hscode": "HS code",
        "multi_col_qty": "Qty",
        "multi_col_fob": "FOB/unit (USD)",
        "multi_col_weight": "Weight/unit (kg)",
        "multi_col_dai": "DAI (%)",
        "multi_logistics_header": "**Logistics (total shipment)**",
        "multi_freight_label": "Sea/air freight total (USD)",
        "multi_insurance_label": "Insurance total (USD)",
        "multi_inland_label": "Inland transport total (PYG)",
        "multi_inland_iva": "Inland transport includes 10 % IVA",
        "multi_customs_header": "**Shared customs fees (PYG)**",
        "multi_valoracion_label": "Customs valuation",
        "multi_val_opt_auto": "0.5 % of CIF per product",
        "multi_val_opt_manual": "Manual total amount",
        "multi_val_manual_label": "Total valuation (PYG)",
        "multi_indi_label": "INDI rate (% of DAI)",
        "multi_canon_label": "Canon SOFIA total (PYG)",
        "multi_consulado_label": "Consular fees total (PYG)",
        "multi_tasa_label": "Port fee total (PYG)",
        "multi_services_header": "**Services & other**",
        "multi_despachante_label": "Broker fee total (PYG)",
        "multi_despachante_iva": "Broker fee includes 10 % IVA",
        "multi_sonstiges_label": "Other costs total (PYG)",
        "multi_alloc_header": "**Allocation keys**",
        "multi_alloc_freight_label": "Freight & insurance allocation",
        "multi_alloc_freight_opt_weight": "Weight‑based (kg share)",
        "multi_alloc_freight_opt_value": "Value‑based (FOB share)",
        "multi_alloc_freight_help": "Freight costs are usually weight‑ or volume‑based.",
        "multi_alloc_local_label": "Port, broker, inland allocation",
        "multi_alloc_local_opt_value": "Value‑based (FOB share)",
        "multi_alloc_local_opt_weight": "Weight‑based (kg share)",
        "multi_alloc_local_help": "How fixed local fees are distributed across products.",
        "multi_warn_no_products": "⚠️ Please add at least one product.",
        "multi_error_zero": "Total FOB value and total weight must be > 0.",
        "multi_results_header": "### 📊 Overall results (multi‑product import)",
        "multi_cap_title": "Total acquisition cost (IAS 2 capitalised)",
        "multi_cap_sub": "Approx. <b>{usd}</b>",
        "multi_credit_title": "Total tax credit (not capitalised)",
        "multi_credit_sub": "Customs IVA ({iva}) + Percepción ({perc}) + Service IVA ({svc})",
        "multi_table_header": "**Result overview per product**",
        "multi_csv_button": "📥 Export allocation sheet as CSV",
        "multi_xlsx_button": "📥 Export allocation sheet as Excel",
        "multi_csv_filename": "import_calculation_multiproduct.csv",
        "multi_xlsx_filename": "import_calculation_multiproduct.xlsx",
        "multi_chart_products": "**Acquisition cost by product**",
        "multi_chart_types": "**Cost breakdown by type**",
        "multi_legend_products": "Products",
        "multi_legend_types": "Cost groups",
        "multi_chart_fob": "FOB goods value",
        "multi_chart_freight": "Int. freight & ins.",
        "multi_chart_customs": "Customs & duties",
        "multi_chart_local": "Local logistics/broker",
        "res_col_product": "Product",
        "res_col_qty": "Qty",
        "res_col_fob": "FOB USD (total)",
        "res_col_weight": "Weight (total)",
        "res_col_cif": "CIF (PYG)",
        "res_col_dai": "DAI duty (PYG)",
        "res_col_other": "Other (PYG)",
        "res_col_total": "Acq. cost (PYG)",
        "res_col_unit_pyg": "Unit cost (PYG)",
        "res_col_unit_usd": "Unit cost (USD)",
        "res_col_credit": "Tax credit (PYG)",
        "footer_header": "### 📚 Glossary & tax foundations (Paraguay)",
    },
    # -----------------------------------------------------------------
    "es": {
        "lang_selector": "🌐 Idioma",
        "zone_label": "Zona de origen",
        "zone_help": "Afecta la tasa DAI y estimaciones de flete segun region de origen",
        "sidebar_global_params": "### 🌍 Parámetros globales",
        "exchange_rate_label": "Tipo de cambio (PYG/USD)",
        "exchange_rate_help": "Tipo de cambio actual Guaraní paraguayo a Dólar estadounidense.",
        "percepcion_ire_label": "Percepción IRE (%)",
        "legal_notice": "📋 Esta herramienta es una ayuda de cálculo. Los resultados no reemplazan asesoría profesional tributaria o aduanera.",
        "reset_button": "🔄 Restablecer todos los campos",
        "banner_title": "🇵🇾 Calculadora de costos de importación – Paraguay",
        "banner_subtitle": "Determinación del costo de adquisición conforme a NIC 2 y Ley 6380/19",
        "tab_single": "📦 Calculadora de producto individual",
        "tab_multi": "🗃️ Calculadora de varios productos",
        "single_card_header": "Datos de cálculo para un producto individual",
        "col1_header": "**1. Especificaciones del producto**",
        "product_name_label": "Nombre del producto",
        "hs_code_label": "Código HS",
        "product_qty_label": "Cantidad (unidades)",
        "fob_price_label": "Precio FOB por unidad (USD)",
        "product_weight_label": "Peso por unidad (kg)",
        "hs_code_invalid": "⚠️ Código HS no válido – ingrese 6‑10 dígitos.",
        "col2_header": "**2. Logística y transporte**",
        "freight_label": "Flete internacional (USD)",
        "insurance_label": "Seguro de transporte (USD)",
        "inland_transport_label": "Transporte interno (PYG)",
        "inland_iva_checkbox": "Transporte interno incluye 10 % IVA",
        "col3_header": "**3. Aduanas y despacho (PYG)**",
        "dai_rate_label": "Arancel DAI (%)",
        "valoracion_label": "Tasa de valoración aduanera",
        "valoracion_opt_auto": "0,5 % del valor CIF",
        "valoracion_opt_manual": "Monto manual",
        "valoracion_manual_label": "Monto manual de valoración (PYG)",
        "indi_rate_label": "Tasa INDI (% del DAI)",
        "canon_sofia_label": "Canon SOFIA (PYG)",
        "consulado_label": "Tarifas consulares (PYG)",
        "tasa_portuaria_label": "Tarifa portuaria (PYG)",
        "despachante_label": "Honorario del agente de aduanas (PYG)",
        "despachante_iva_checkbox": "Honorario incluye 10 % IVA",
        "other_costs_label": "Otros costos (PYG)",
        "results_header": "### 📊 Resultados del cálculo (producto individual)",
        "unit_cost_title": "Costo unitario de adquisición (NIC 2 capitalizado)",
        "unit_cost_sub": "Equivale a <b>{usd}</b> por unidad (costo total: {total})",
        "tax_credit_title": "Crédito fiscal (no capitalizado)",
        "tax_credit_sub": "Se compensa directamente con la declaración de impuestos.",
        "multi_card_header": "Cálculo de importación multiproducto con prorrateo",
        "multi_intro": "Ingrese su lista de productos. Los costos compartidos se prorratean por valor (FOB) o peso (kg).",
        "multi_products_header": "**1. Productos importados**",
        "multi_col_product": "Producto",
        "multi_col_hscode": "Código HS",
        "multi_col_qty": "Cant.",
        "multi_col_fob": "FOB/und. (USD)",
        "multi_col_weight": "Peso/und. (kg)",
        "multi_col_dai": "DAI (%)",
        "multi_logistics_header": "**Logística (envío total)**",
        "multi_freight_label": "Flete marítimo/aéreo total (USD)",
        "multi_insurance_label": "Seguro total (USD)",
        "multi_inland_label": "Transporte interno total (PYG)",
        "multi_inland_iva": "Transporte interno incluye 10 % IVA",
        "multi_customs_header": "**Tasas aduaneras compartidas (PYG)**",
        "multi_valoracion_label": "Valoración aduanera",
        "multi_val_opt_auto": "0,5 % del CIF por producto",
        "multi_val_opt_manual": "Monto total manual",
        "multi_val_manual_label": "Valoración total (PYG)",
        "multi_indi_label": "Tasa INDI (% del DAI)",
        "multi_canon_label": "Canon SOFIA total (PYG)",
        "multi_consulado_label": "Tarifas consulares total (PYG)",
        "multi_tasa_label": "Tarifa portuaria total (PYG)",
        "multi_services_header": "**Servicios y otros**",
        "multi_despachante_label": "Honorario despachante total (PYG)",
        "multi_despachante_iva": "Honorario incluye 10 % IVA",
        "multi_sonstiges_label": "Otros costos total (PYG)",
        "multi_alloc_header": "**Claves de prorrateo**",
        "multi_alloc_freight_label": "Prorrateo de flete y seguro",
        "multi_alloc_freight_opt_weight": "Por peso (kg)",
        "multi_alloc_freight_opt_value": "Por valor (FOB)",
        "multi_alloc_freight_help": "El flete suele asignarse por peso o volumen.",
        "multi_alloc_local_label": "Prorrateo puerto, broker, interno",
        "multi_alloc_local_opt_value": "Por valor (FOB)",
        "multi_alloc_local_opt_weight": "Por peso (kg)",
        "multi_alloc_local_help": "Cómo se distribuyen las tarifas locales fijas.",
        "multi_warn_no_products": "⚠️ Agregue al menos un producto.",
        "multi_error_zero": "El valor FOB total y el peso total deben ser > 0.",
        "multi_results_header": "### 📊 Resultados generales (importación multiproducto)",
        "multi_cap_title": "Costo de adquisición total (NIC 2 capitalizado)",
        "multi_cap_sub": "Aprox. <b>{usd}</b>",
        "multi_credit_title": "Crédito fiscal total (no capitalizado)",
        "multi_credit_sub": "IVA aduanero ({iva}) + Percepción ({perc}) + IVA servicios ({svc})",
        "multi_table_header": "**Resumen por producto**",
        "multi_csv_button": "📥 Exportar prorrateo como CSV",
        "multi_xlsx_button": "📥 Exportar prorrateo como Excel",
        "multi_csv_filename": "calculo_importacion_multiproducto.csv",
        "multi_xlsx_filename": "calculo_importacion_multiproducto.xlsx",
        "multi_chart_products": "**Costo de adquisición por producto**",
        "multi_chart_types": "**Desglose por tipo de costo**",
        "multi_legend_products": "Productos",
        "multi_legend_types": "Grupos de costo",
        "multi_chart_fob": "Valor FOB",
        "multi_chart_freight": "Flete & seguro int.",
        "multi_chart_customs": "Aranceles y tasas",
        "multi_chart_local": "Logística local/broker",
        "res_col_product": "Producto",
        "res_col_qty": "Cant.",
        "res_col_fob": "FOB USD (total)",
        "res_col_weight": "Peso (total)",
        "res_col_cif": "CIF (PYG)",
        "res_col_dai": "DAI (PYG)",
        "res_col_other": "Otros (PYG)",
        "res_col_total": "Costo adq. (PYG)",
        "res_col_unit_pyg": "Costo unit. (PYG)",
        "res_col_unit_usd": "Costo unit. (USD)",
        "res_col_credit": "Crédito fiscal (PYG)",
        "footer_header": "### 📚 Glosario y fundamentos tributarios (Paraguay)",
    },
}


def t(key: str) -> str:
    """Return translated string for *key* using the language stored in
    ``st.session_state['lang']``.  Falls back to German, then returns the key
    itself if not found."""
    lang = st.session_state.get("lang", "de")
    tr = TRANSLATIONS.get(lang, TRANSLATIONS["de"])
    return tr.get(key, TRANSLATIONS["de"].get(key, key))


# =====================================================================
# HS‑CODE  –  Authority warnings (Paraguay)
# =====================================================================
# Each rule: (list of 2‑digit prefixes, authority, {lang: warning_text})
HS_CODE_RULES: list[dict] = [
    {
        "prefixes": ["01", "02", "03", "04", "05"],
        "authority": "SENACSA",
        "icon": "🔴",
        "warning": {
            "de": (
                "Tierische Produkte – Genehmigung des SENACSA "
                "(Servicio Nacional de Calidad y Salud Animal) erforderlich. "
                "Es fallen Inspektionsgebühren und Gesundheitszertifikate an. "
                "Sondergenehmigungen können nötig sein."
            ),
            "en": (
                "Animal products – SENACSA (National Animal Quality & Health Service) "
                "approval required. Inspection fees and sanitary certificates apply. "
                "Special permits may be needed."
            ),
            "es": (
                "Productos de origen animal – Se requiere aprobación del SENACSA "
                "(Servicio Nacional de Calidad y Salud Animal). "
                "Aplican tasas de inspección y certificados sanitarios. "
                "Pueden requerirse permisos especiales."
            ),
        },
    },
    {
        "prefixes": ["06", "07", "08", "09", "10", "11", "12", "13", "14"],
        "authority": "SENAVE",
        "icon": "🔴",
        "warning": {
            "de": (
                "Pflanzliche Produkte – Genehmigung des SENAVE "
                "(Servicio Nacional de Calidad Vegetal y de Semillas) erforderlich. "
                "Phytosanitäre Inspektionen und Zertifikate können Zusatzkosten verursachen."
            ),
            "en": (
                "Plant products – SENAVE (National Plant Quality & Seed Service) "
                "approval required. Phytosanitary inspections and certificates "
                "may incur additional costs."
            ),
            "es": (
                "Productos vegetales – Se requiere aprobación del SENAVE "
                "(Servicio Nacional de Calidad Vegetal y de Semillas). "
                "Las inspecciones fitosanitarias y certificados pueden generar costos adicionales."
            ),
        },
    },
    {
        "prefixes": ["15"],
        "authority": "SENACSA / SENAVE",
        "icon": "🔴",
        "warning": {
            "de": (
                "Fette & Öle – Je nach Ursprung (tierisch/pflanzlich) ist SENACSA "
                "oder SENAVE zuständig. Zusätzliche Inspektionsgebühren möglich."
            ),
            "en": (
                "Fats & oils – Depending on origin (animal/plant), SENACSA or SENAVE "
                "has jurisdiction. Additional inspection fees may apply."
            ),
            "es": (
                "Grasas y aceites – Según el origen (animal/vegetal), SENACSA o SENAVE "
                "tiene jurisdicción. Pueden aplicarse tarifas de inspección adicionales."
            ),
        },
    },
    {
        "prefixes": ["16", "17", "18", "19", "20", "21"],
        "authority": "INAN",
        "icon": "🔴",
        "warning": {
            "de": (
                "Verarbeitete Lebensmittel – INAN (Instituto Nacional de Alimentación "
                "y Nutrición) ist zuständig. Es sind Registro Sanitario, "
                "Laboranalysen und Etikettiergenehmigungen erforderlich. "
                "Zusätzliche Gebühren fallen an."
            ),
            "en": (
                "Processed food – INAN (National Food & Nutrition Institute) "
                "has jurisdiction. Sanitary registration, laboratory analyses, "
                "and labelling approvals are required. Additional fees apply."
            ),
            "es": (
                "Alimentos procesados – INAN (Instituto Nacional de Alimentación "
                "y Nutrición) tiene jurisdicción. Se requieren Registro Sanitario, "
                "análisis de laboratorio y aprobaciones de etiquetado. "
                "Aplican tarifas adicionales."
            ),
        },
    },
    {
        "prefixes": ["22"],
        "authority": "INAN + Impuestos Selectivos",
        "icon": "🔴",
        "warning": {
            "de": (
                "Getränke & Alkohol – Genehmigung des INAN erforderlich. "
                "Auf alkoholische Getränke fällt zusätzlich ein Impuesto Selectivo "
                "al Consumo (ISC) an. Sondergenehmigungen für Spirituosen nötig."
            ),
            "en": (
                "Beverages & alcohol – INAN approval required. "
                "Alcoholic beverages are additionally subject to Selective "
                "Consumption Tax (ISC). Special permits needed for spirits."
            ),
            "es": (
                "Bebidas y alcohol – Se requiere aprobación del INAN. "
                "Las bebidas alcohólicas están sujetas además al Impuesto Selectivo "
                "al Consumo (ISC). Se necesitan permisos especiales para licores."
            ),
        },
    },
    {
        "prefixes": ["24"],
        "authority": "ISC / Aduanas",
        "icon": "🔴",
        "warning": {
            "de": (
                "Tabakwaren – Es fällt der Impuesto Selectivo al Consumo (ISC) an. "
                "Spezielle Einfuhrgenehmigungen und Gesundheitswarnungen "
                "auf der Verpackung sind obligatorisch."
            ),
            "en": (
                "Tobacco products – Subject to Selective Consumption Tax (ISC). "
                "Special import permits and health warnings on packaging are mandatory."
            ),
            "es": (
                "Productos de tabaco – Sujetos al Impuesto Selectivo al Consumo (ISC). "
                "Se requieren permisos especiales de importación y "
                "advertencias sanitarias obligatorias en el empaque."
            ),
        },
    },
    {
        "prefixes": ["28", "29", "31", "32", "33", "34", "35", "37", "38"],
        "authority": "MADES",
        "icon": "🟡",
        "warning": {
            "de": (
                "Chemische Produkte – Genehmigung des MADES "
                "(Ministerio del Ambiente y Desarrollo Sostenible) kann erforderlich sein. "
                "Umweltverträglichkeitsprüfungen und Gefahrgut‑Dokumentation "
                "können Zusatzkosten verursachen."
            ),
            "en": (
                "Chemical products – MADES (Ministry of Environment) "
                "approval may be required. Environmental impact assessments "
                "and hazardous goods documentation may incur additional costs."
            ),
            "es": (
                "Productos químicos – Puede requerirse aprobación del MADES "
                "(Ministerio del Ambiente y Desarrollo Sostenible). "
                "Las evaluaciones de impacto ambiental y documentación de "
                "mercancías peligrosas pueden generar costos adicionales."
            ),
        },
    },
    {
        "prefixes": ["30"],
        "authority": "DINAVISA",
        "icon": "🔴",
        "warning": {
            "de": (
                "Pharmazeutische Produkte – Genehmigung der DINAVISA "
                "(Dirección Nacional de Vigilancia Sanitaria) zwingend erforderlich. "
                "Registro Sanitario, Laborprüfungen und Lagerungsauflagen "
                "verursachen erhebliche Zusatzkosten."
            ),
            "en": (
                "Pharmaceutical products – DINAVISA (National Health Surveillance "
                "Directorate) approval mandatory. Sanitary registration, "
                "laboratory testing, and storage requirements incur significant "
                "additional costs."
            ),
            "es": (
                "Productos farmacéuticos – Se requiere obligatoriamente la aprobación "
                "de la DINAVISA (Dirección Nacional de Vigilancia Sanitaria). "
                "El Registro Sanitario, pruebas de laboratorio y requisitos de "
                "almacenamiento generan costos adicionales significativos."
            ),
        },
    },
    {
        "prefixes": ["36", "93"],
        "authority": "DIMABEL",
        "icon": "🔴",
        "warning": {
            "de": (
                "Waffen, Munition oder Explosivstoffe – Genehmigung der DIMABEL "
                "(Dirección de Material Bélico) zwingend erforderlich. "
                "Strenge Einfuhrbeschränkungen und erhebliche Zusatzgebühren."
            ),
            "en": (
                "Weapons, ammunition, or explosives – DIMABEL (Military "
                "Materials Directorate) approval mandatory. Strict import "
                "restrictions and significant additional fees."
            ),
            "es": (
                "Armas, municiones o explosivos – Se requiere obligatoriamente "
                "la aprobación de la DIMABEL (Dirección de Material Bélico). "
                "Restricciones estrictas de importación y tarifas adicionales significativas."
            ),
        },
    },
    {
        "prefixes": ["84", "85"],
        "authority": "DNA / CONATEL",
        "icon": "🟡",
        "warning": {
            "de": (
                "Maschinen & Elektronik – Bei bestimmten Geräten kann eine "
                "Genehmigung der CONATEL (Telekommunikation) oder Anti‑Dumping‑"
                "Zölle anfallen. Prüfen Sie die spezifische Tarifposition."
            ),
            "en": (
                "Machinery & electronics – Certain devices may require CONATEL "
                "(telecommunications) approval or be subject to anti‑dumping "
                "duties. Verify the specific tariff heading."
            ),
            "es": (
                "Maquinarias y electrónicos – Ciertos dispositivos pueden requerir "
                "aprobación de CONATEL (telecomunicaciones) o estar sujetos a "
                "derechos anti‑dumping. Verifique la posición arancelaria específica."
            ),
        },
    },
    {
        "prefixes": ["87"],
        "authority": "DINATRAN / MIC",
        "icon": "🟡",
        "warning": {
            "de": (
                "Fahrzeuge – Zusätzliche Emissionsprüfungen und Homologation "
                "durch das MIC können erforderlich sein. Der ISC (Impuesto "
                "Selectivo) kann anfallen."
            ),
            "en": (
                "Vehicles – Additional emission checks and MIC homologation "
                "may be required. Selective Consumption Tax (ISC) may apply."
            ),
            "es": (
                "Vehículos – Pueden requerirse controles de emisiones adicionales "
                "y homologación del MIC. Puede aplicarse el Impuesto Selectivo "
                "al Consumo (ISC)."
            ),
        },
    },
]


def check_hs_code(code: str) -> dict | None:
    """Check *code* against HS‑code rules.

    Returns a dict ``{'authority', 'icon', 'warning'}`` or ``None``.
    The warning text is already in the current UI language.
    """
    if not code or not code.strip():
        return None
    code = code.strip()

    # Must be digits only for a valid check
    if not re.fullmatch(r"\d{6,10}", code):
        return None

    lang = st.session_state.get("lang", "de")
    prefix2 = code[:2]

    for rule in HS_CODE_RULES:
        if prefix2 in rule["prefixes"]:
            return {
                "authority": rule["authority"],
                "icon": rule.get("icon", "🔴"),
                "warning": rule["warning"].get(lang, rule["warning"]["en"]),
            }
    return None


# =====================================================================
# STATE PERSISTENCE  –  save / load / reset
# =====================================================================

def load_state() -> None:
    """Load persisted state from JSON file into ``st.session_state``.
    Only sets keys that are NOT yet in session_state (so widget values
    from the current session always win)."""
    if not os.path.exists(STATE_FILE):
        return
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        for key, val in data.items():
            if key not in st.session_state:
                st.session_state[key] = val
    except Exception:
        pass  # corrupt file – ignore silently


def save_state() -> None:
    """Persist the current session_state scalars to a JSON file."""
    state: dict = {}
    for key in STATE_KEYS:
        if key in st.session_state:
            state[key] = st.session_state[key]
    # Multi‑product DataFrame (optional)
    if "df_products" in st.session_state:
        import pandas as pd
        df = st.session_state["df_products"]
        if isinstance(df, pd.DataFrame):
            state["_df_products_records"] = df.to_dict(orient="records")
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as fh:
            json.dump(state, fh, ensure_ascii=False)
    except Exception:
        pass


def reset_state() -> None:
    """Callback for the reset button. Sets all scalar keys back to defaults
    and removes the persisted state file. Does NOT touch the product table."""
    for key, val in DEFAULTS.items():
        st.session_state[key] = val
    if os.path.exists(STATE_FILE):
        try:
            os.remove(STATE_FILE)
        except Exception:
            pass


def init_defaults() -> None:
    """Ensure every expected key exists in session_state, using the
    default value if not already present (from persistence or widget)."""
    for key, val in DEFAULTS.items():
        if key not in st.session_state:
            st.session_state[key] = val


# =====================================================================
# DATABASE  –  Product HS‑Code Lookup (products_hs.db)
# =====================================================================

import sqlite3

def get_db_path() -> str:
    """Return absolute path to the products database, looking in the
    ``data/`` subdirectory alongside this file."""
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "data", "products_hs.db")


def _get_conn():
    """Open (or reuse) a DB connection. Always returns a new one — the
    caller is responsible for closing it."""
    db_path = get_db_path()
    if not os.path.exists(db_path):
        return None
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def lookup_hs_product(hs_code: str) -> dict | None:
    """Look up a product by exact HS code match.

    Returns a dict with keys: description, hs_code, category,
    default_dai, typical_fob_usd, typical_weight_kg.
    Returns None if not found or DB missing.
    """
    conn = _get_conn()
    if conn is None:
        return None
    try:
        cur = conn.execute(
            "SELECT description, hs_code, category, default_dai, typical_fob_usd, typical_weight_kg "
            "FROM products WHERE hs_code = ? LIMIT 1",
            (hs_code.strip(),),
        )
        row = cur.fetchone()
        if row:
            return dict(row)
        return None
    except Exception:
        return None
    finally:
        conn.close()


def search_products(query: str, limit: int = 20) -> list[dict]:
    """Full‑text search across product descriptions and HS codes.

    Returns a list of dicts, each with: id, description, hs_code,
    category, default_dai, typical_fob_usd, typical_weight_kg.
    """
    conn = _get_conn()
    if conn is None:
        return []
    try:
        like = f"%{query}%"
        cur = conn.execute(
            "SELECT id, description, hs_code, category, default_dai, typical_fob_usd, typical_weight_kg "
            "FROM products WHERE description LIKE ? OR hs_code LIKE ? "
            "ORDER BY description LIMIT ?",
            (like, like, limit),
        )
        return [dict(row) for row in cur.fetchall()]
    except Exception:
        return []
    finally:
        conn.close()


def list_all_products(limit: int = 200) -> list[dict]:
    """Return all products from the database, ordered by description."""
    conn = _get_conn()
    if conn is None:
        return []
    try:
        cur = conn.execute(
            "SELECT description, hs_code, category, default_dai, typical_fob_usd, typical_weight_kg "
            "FROM products ORDER BY description LIMIT ?",
            (limit,),
        )
        return [dict(row) for row in cur.fetchall()]
    except Exception:
        return []
    finally:
        conn.close()


def get_product_categories() -> list[str]:
    """Return distinct product categories from the database."""
    conn = _get_conn()
    if conn is None:
        return []
    try:
        cur = conn.execute(
            "SELECT DISTINCT category FROM products WHERE category IS NOT NULL ORDER BY category"
        )
        return [row[0] for row in cur.fetchall()]
    except Exception:
        return []
    finally:
        conn.close()


# =====================================================================
# CALCULATION ENGINE  –  pure functions
# .. deprecated:: Verwende stattdessen calculator.py
# =====================================================================

def calc_single_product(params: dict, ex_rate: float, percep_ire_rate: float) -> dict:
    import warnings
    warnings.warn("Verwende stattdessen calculator.calc_single_product", DeprecationWarning, stacklevel=2)
    """Calculate single-product import costs.

    Args:
        params: dict with keys matching DEFAULTS (p_name, hs_code, p_qty,
                p_fob_usd, p_weight, freight_usd, insurance_usd, inland_pyg,
                inland_iva_incl, dai_rate, val_mode, val_pyg_input, indi_rate,
                canon_sofia, consulado, tasa_portuaria, despachante,
                despachante_iva_incl, sonstiges)
        ex_rate: PYG/USD exchange rate
        percep_ire_rate: IRE perception rate in percent

    Returns:
        dict with all intermediate and final values
    """
    p_qty = params.get("p_qty", 0)
    p_fob_usd = params.get("p_fob_usd", 0.0)
    freight_usd = params.get("freight_usd", 0.0)
    insurance_usd = params.get("insurance_usd", 0.0)
    inland_pyg = params.get("inland_pyg", 0.0)
    inland_iva_incl = params.get("inland_iva_incl", True)
    dai_rate = params.get("dai_rate", 0.0)
    val_mode = params.get("val_mode", 0)
    val_pyg_input = params.get("val_pyg_input", 0.0)
    indi_rate = params.get("indi_rate", 0.0)
    canon_sofia = params.get("canon_sofia", 0.0)
    consulado = params.get("consulado", 0.0)
    tasa_portuaria = params.get("tasa_portuaria", 0.0)
    despachante = params.get("despachante", 0.0)
    despachante_iva_incl = params.get("despachante_iva_incl", True)
    sonstiges = params.get("sonstiges", 0.0)

    total_fob_usd = p_qty * p_fob_usd
    total_fob_pyg = total_fob_usd * ex_rate
    total_freight_pyg = freight_usd * ex_rate
    total_insurance_pyg = insurance_usd * ex_rate

    cif_usd = total_fob_usd + freight_usd + insurance_usd
    cif_pyg = cif_usd * ex_rate

    # Netting inland transport
    if inland_iva_incl:
        inland_netto = inland_pyg / 1.1
        inland_iva = inland_netto * 0.1
    else:
        inland_netto = inland_pyg
        inland_iva = 0.0

    # Netting despachante
    if despachante_iva_incl:
        despachante_netto = despachante / 1.1
        despachante_iva = despachante_netto * 0.1
    else:
        despachante_netto = despachante
        despachante_iva = 0.0

    # Customs taxes
    dai_pyg = cif_pyg * (dai_rate / 100.0)

    if val_mode == 0:
        val_pyg = cif_pyg * 0.005
    else:
        val_pyg = val_pyg_input

    indi_pyg = dai_pyg * (indi_rate / 100.0)

    # Capitalized cost (IAS 2)
    capitalized_logistics = total_freight_pyg + total_insurance_pyg + inland_netto
    capitalized_customs = (
        dai_pyg + val_pyg + indi_pyg
        + canon_sofia + consulado + tasa_portuaria + despachante_netto + sonstiges
    )
    total_acquisition_cost = total_fob_pyg + capitalized_logistics + capitalized_customs

    if p_qty > 0:
        unit_cost_pyg = total_acquisition_cost / p_qty
        unit_cost_usd = unit_cost_pyg / ex_rate
    else:
        unit_cost_pyg = 0.0
        unit_cost_usd = 0.0

    # Tax credits (Ley 6380/19) – NOT capitalised
    base_iva_aduana = (
        cif_pyg + dai_pyg + val_pyg + indi_pyg
        + canon_sofia + consulado + tasa_portuaria
    )
    iva_importacion = base_iva_aduana * 0.10
    percepcion_ire = cif_pyg * (percep_ire_rate / 100.0)
    total_tax_credit = iva_importacion + percepcion_ire + inland_iva + despachante_iva

    return {
        "total_fob_usd": total_fob_usd,
        "total_fob_pyg": total_fob_pyg,
        "cif_usd": cif_usd,
        "cif_pyg": cif_pyg,
        "inland_netto": inland_netto,
        "inland_iva": inland_iva,
        "despachante_netto": despachante_netto,
        "despachante_iva": despachante_iva,
        "dai_pyg": dai_pyg,
        "val_pyg": val_pyg,
        "indi_pyg": indi_pyg,
        "capitalized_logistics": capitalized_logistics,
        "capitalized_customs": capitalized_customs,
        "total_acquisition_cost": total_acquisition_cost,
        "unit_cost_pyg": unit_cost_pyg,
        "unit_cost_usd": unit_cost_usd,
        "iva_importacion": iva_importacion,
        "percepcion_ire": percepcion_ire,
        "total_tax_credit": total_tax_credit,
        "base_iva_aduana": base_iva_aduana,
    }


def calc_multi_product(
    products_df,
    ex_rate: float,
    percep_ire_rate: float,
    multi_freight_usd: float,
    multi_insurance_usd: float,
    multi_inland_pyg: float,
    multi_inland_iva_incl: bool,
    multi_val_mode: int,
    multi_val_pyg_manual: float,
    multi_indi_rate: float,
    multi_canon_sofia: float,
    multi_consulado: float,
    multi_tasa_portuaria: float,
    multi_despachante: float,
    multi_despachante_iva_incl: bool,
    multi_sonstiges: float,
    alloc_freight: int,  # 0=weight, 1=value
    alloc_local: int,    # 0=value, 1=weight
) -> dict:
    import warnings
    warnings.warn("Verwende stattdessen calculator.calc_multi_product", DeprecationWarning, stacklevel=2)
    """Calculate multi-product import costs.

    Returns:
        dict with keys: products_df (enriched DataFrame), total_capitalized,
        total_tax_credit, inland_iva, desp_iva, summary dicts
    """
    import pandas as pd

    if len(products_df) == 0:
        return None

    # Netting services
    if multi_inland_iva_incl:
        m_inland_netto = multi_inland_pyg / 1.1
        m_inland_iva = m_inland_netto * 0.1
    else:
        m_inland_netto = multi_inland_pyg
        m_inland_iva = 0.0

    if multi_despachante_iva_incl:
        m_desp_netto = multi_despachante / 1.1
        m_desp_iva = m_desp_netto * 0.1
    else:
        m_desp_netto = multi_despachante
        m_desp_iva = 0.0

    prods = products_df.copy()
    prods["Produktname"] = prods["Produktname"].fillna("-")
    prods["Menge"] = pd.to_numeric(prods["Menge"], errors="coerce").fillna(0).astype(int)
    prods["FOB pro Stk. (USD)"] = pd.to_numeric(prods["FOB pro Stk. (USD)"], errors="coerce").fillna(0.0)
    prods["Gewicht pro Stk. (kg)"] = pd.to_numeric(prods["Gewicht pro Stk. (kg)"], errors="coerce").fillna(0.0)
    prods["DAI (%)"] = pd.to_numeric(prods["DAI (%)"], errors="coerce").fillna(0.0)

    prods["Total_FOB_USD"] = prods["Menge"] * prods["FOB pro Stk. (USD)"]
    prods["Total_Weight_kg"] = prods["Menge"] * prods["Gewicht pro Stk. (kg)"]

    sum_fob_usd = prods["Total_FOB_USD"].sum()
    sum_weight_kg = prods["Total_Weight_kg"].sum()

    if sum_fob_usd <= 0 or sum_weight_kg <= 0:
        return None

    prods["FOB_Share"] = prods["Total_FOB_USD"] / sum_fob_usd
    prods["Weight_Share"] = prods["Total_Weight_kg"] / sum_weight_kg

    # Freight & Insurance allocation
    use_weight_freight = (alloc_freight == 0)
    share_freight = prods["Weight_Share"] if use_weight_freight else prods["FOB_Share"]
    prods["Alloc_Freight_USD"] = multi_freight_usd * share_freight
    prods["Alloc_Insurance_USD"] = multi_insurance_usd * share_freight
    prods["Alloc_Freight_PYG"] = prods["Alloc_Freight_USD"] * ex_rate
    prods["Alloc_Insurance_PYG"] = prods["Alloc_Insurance_USD"] * ex_rate

    # CIF
    prods["CIF_USD"] = prods["Total_FOB_USD"] + prods["Alloc_Freight_USD"] + prods["Alloc_Insurance_USD"]
    prods["CIF_PYG"] = prods["CIF_USD"] * ex_rate

    # Product-specific duties
    prods["DAI_PYG"] = prods["CIF_PYG"] * (prods["DAI (%)"] / 100.0)

    if multi_val_mode == 0:
        prods["Val_PYG"] = prods["CIF_PYG"] * 0.005
    else:
        use_weight_local = (alloc_local == 1)
        share_local_val = prods["Weight_Share"] if use_weight_local else prods["FOB_Share"]
        prods["Val_PYG"] = multi_val_pyg_manual * share_local_val

    prods["INDI_PYG"] = prods["DAI_PYG"] * (multi_indi_rate / 100.0)

    # Allocate common local costs
    total_common_local = (
        multi_canon_sofia + multi_consulado + multi_tasa_portuaria
        + m_desp_netto + multi_sonstiges + m_inland_netto
    )
    use_weight_local = (alloc_local == 1)
    share_local = prods["Weight_Share"] if use_weight_local else prods["FOB_Share"]
    prods["Alloc_Local_PYG"] = total_common_local * share_local
    prods["Alloc_TaxBase_Fees_PYG"] = (
        multi_canon_sofia + multi_consulado + multi_tasa_portuaria
    ) * share_local

    # Total capitalised cost
    prods["Total_Capitalized_PYG"] = (
        prods["Total_FOB_USD"] * ex_rate
        + prods["Alloc_Freight_PYG"]
        + prods["Alloc_Insurance_PYG"]
        + prods["DAI_PYG"]
        + prods["Val_PYG"]
        + prods["INDI_PYG"]
        + prods["Alloc_Local_PYG"]
    )
    prods["Stückkosten_PYG"] = prods.apply(
        lambda r: r["Total_Capitalized_PYG"] / r["Menge"] if r["Menge"] > 0 else 0, axis=1
    )
    prods["Stückkosten_USD"] = prods["Stückkosten_PYG"] / ex_rate

    # Tax credits
    prods["IVA_Base_PYG"] = (
        prods["CIF_PYG"] + prods["DAI_PYG"] + prods["Val_PYG"]
        + prods["INDI_PYG"] + prods["Alloc_TaxBase_Fees_PYG"]
    )
    prods["IVA_Importacion_PYG"] = prods["IVA_Base_PYG"] * 0.10
    prods["Percepcion_IRE_PYG"] = prods["CIF_PYG"] * (percep_ire_rate / 100.0)

    prods["Alloc_Inland_IVA"] = m_inland_iva * share_local
    prods["Alloc_Broker_IVA"] = m_desp_iva * share_local

    prods["Total_Tax_Credit_PYG"] = (
        prods["IVA_Importacion_PYG"] + prods["Percepcion_IRE_PYG"]
        + prods["Alloc_Inland_IVA"] + prods["Alloc_Broker_IVA"]
    )

    total_shipment_cap = prods["Total_Capitalized_PYG"].sum()
    total_shipment_credit = prods["Total_Tax_Credit_PYG"].sum()

    return {
        "products_df": prods,
        "total_capitalized": total_shipment_cap,
        "total_tax_credit": total_shipment_credit,
        "inland_iva": m_inland_iva,
        "desp_iva": m_desp_iva,
        "sum_fob_usd": sum_fob_usd,
        "sum_weight_kg": sum_weight_kg,
    }

#!/usr/bin/env python3
import json, os, sqlite3, urllib.request

# =====================================================================
# USAGE TRACKING
# =====================================================================
def get_usage_db_path():
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "data", "usage.db")

def init_usage_db():
    db_path = get_usage_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""CREATE TABLE IF NOT EXISTS sessions (
        id TEXT PRIMARY KEY, started_at TEXT NOT NULL DEFAULT (datetime('now')),
        last_active TEXT NOT NULL DEFAULT (datetime('now')),
        ip_address TEXT, user_agent TEXT, total_calculations INTEGER DEFAULT 0)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS calculations (
        id INTEGER PRIMARY KEY AUTOINCREMENT, session_id TEXT NOT NULL,
        calc_at TEXT NOT NULL DEFAULT (datetime('now')),
        product_name TEXT, hs_code TEXT, hs_source TEXT DEFAULT 'manual',
        quantity REAL, fob_usd REAL, weight_kg REAL, dai_pct REAL,
        cif_pyg REAL, unit_cost_pyg REAL, total_cost_pyg REAL, tax_credit_pyg REAL)""")
    conn.commit()
    conn.close()

def log_session(session_id, ip_address="", user_agent=""):
    db_path = get_usage_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("INSERT OR REPLACE INTO sessions (id, last_active, ip_address, user_agent) VALUES (?, datetime('now'), ?, ?)", (session_id, ip_address, user_agent[:500]))
    conn.commit()
    conn.close()

def log_calculation(session_id, product_name="", hs_code="", hs_source="manual", quantity=0, fob_usd=0, weight_kg=0, dai_pct=0, cif_pyg=0, unit_cost_pyg=0, total_cost_pyg=0, tax_credit_pyg=0):
    db_path = get_usage_db_path()
    conn = sqlite3.connect(db_path)
    conn.execute("""INSERT INTO calculations (session_id, product_name, hs_code, hs_source, quantity, fob_usd, weight_kg, dai_pct, cif_pyg, unit_cost_pyg, total_cost_pyg, tax_credit_pyg)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""", (session_id, product_name, hs_code, hs_source, quantity, fob_usd, weight_kg, dai_pct, cif_pyg, unit_cost_pyg, total_cost_pyg, tax_credit_pyg))
    conn.execute("UPDATE sessions SET total_calculations = total_calculations + 1, last_active = datetime('now') WHERE id = ?", (session_id,))
    conn.commit()
    conn.close()

def get_admin_stats():
    db_path = get_usage_db_path()
    if not os.path.exists(db_path): return {"sessions": 0, "calculations": 0, "recent": []}
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    sessions = conn.execute("SELECT COUNT(*) as n FROM sessions").fetchone()["n"]
    calcs = conn.execute("SELECT COUNT(*) as n FROM calculations").fetchone()["n"]
    recent = [dict(r) for r in conn.execute("SELECT * FROM calculations ORDER BY calc_at DESC LIMIT 50").fetchall()]
    conn.close()
    return {"sessions": sessions, "calculations": calcs, "recent": recent}

def get_config_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")

def load_config():
    p = get_config_path()
    if os.path.exists(p):
        try:
            with open(p) as f: return json.load(f)
        except: pass
    return {}

def save_config(cfg):
    with open(get_config_path(), "w") as f: json.dump(cfg, f, indent=2)

# =====================================================================
# AI HS-CODE LOOKUP
# =====================================================================
def lookup_hs_with_ai(description, config):
    provider = config.get("ai_provider", "")
    if provider == "ollama": return _ask_ollama(description, config)
    elif provider == "claude": return _ask_claude(description, config)
    return None

def _ask_ollama(description, config):
    model = config.get("ollama_model", "qwen3-coder:480b")
    endpoint = config.get("ollama_endpoint", "https://ollama.com/v1")
    api_key = config.get("ollama_api_key", config.get("claude_api_key", ""))
    prompt = 'You are a customs tariff expert. Return ONLY valid JSON with keys hs_code (8-digit HS code) and explanation (one-line German description). Product: "' + description + '"'
    try:
        data = json.dumps({
            "model": model,
            "messages": [
                {"role": "system", "content": "You are a customs tariff expert. Always respond with valid JSON only."},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.1,
            "max_tokens": 200,
            "response_format": {"type": "json_object"}
        }).encode()
        req = urllib.request.Request(
            endpoint + "/chat/completions",
            data=data,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer " + (api_key or "")
            }
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
            content = data["choices"][0]["message"]["content"]
            response = json.loads(content)
            return {
                "hs_code": response.get("hs_code", ""),
                "explanation": response.get("explanation", "KI-Vorschlag"),
                "tokens_used": data.get("usage", {}).get("total_tokens", 0),
                "provider": "ollama"
            }
    except Exception as e:
        return {"hs_code": "", "explanation": "Fehler: " + str(e)[:100], "tokens_used": 0, "provider": "ollama"}

def _ask_claude(description, config):
    api_key = config.get("claude_api_key", "")
    if not api_key: return None
    model = config.get("claude_model", "claude-3-haiku-20240307")
    prompt = 'Return ONLY valid JSON: {"hs_code":"8-digit","explanation":"German"}. Product: "' + description + '"'
    try:
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=json.dumps({"model": model, "max_tokens": 200, "messages": [{"role": "user", "content": prompt}]}).encode(),
            headers={"Content-Type": "application/json", "x-api-key": api_key, "anthropic-version": "2023-06-01"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
            content = data["content"][0]["text"]
            response = json.loads(content)
            return {"hs_code": response.get("hs_code",""), "explanation": response.get("explanation","KI-Vorschlag"), "tokens_used": data.get("usage",{}).get("input_tokens",0)+data.get("usage",{}).get("output_tokens",0), "provider": "claude"}
    except Exception as e:
        return {"hs_code": "", "explanation": "Fehler: "+str(e)[:100], "tokens_used": 0, "provider": "claude"}

def track_ai_tokens(tokens, config):
    config["ai_tokens_used"] = config.get("ai_tokens_used", 0) + tokens
    save_config(config)

print("EXTENSIONS_LOADED")
def test_ai_connection(config: dict) -> dict:
    provider = config.get('ai_provider', '')
    if not provider:
        return {'ok': False, 'error': 'Kein Provider konfiguriert'}
    if provider == 'ollama':
        return _test_ollama(config)
    elif provider == 'claude':
        return _test_claude(config)
    return {'ok': False, 'error': f'Unbekannter Provider: {provider}'}

def _test_ollama(config):
    import urllib.request, json
    ep = config.get('ollama_endpoint', 'https://ollama.com/v1')
    model = config.get('ollama_model', 'qwen3-coder:480b')
    api_key = config.get('ollama_api_key', config.get('claude_api_key', ''))
    try:
        req = urllib.request.Request(
            ep + '/models',
            headers={'Content-Type': 'application/json', 'Authorization': 'Bearer ' + (api_key or '')}
        )
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
            models = [m['id'] for m in data.get('data', [])]
            return {'ok': True, 'models': models[:10], 'endpoint': ep}
    except Exception as e:
        return {'ok': False, 'error': str(e)[:200], 'endpoint': ep}

def _test_claude(config):
    import urllib.request, json
    api_key = config.get('claude_api_key', '')
    if not api_key:
        return {'ok': False, 'error': 'Kein Claude API Key'}
    try:
        req = urllib.request.Request(
            'https://api.anthropic.com/v1/messages',
            data=json.dumps({'model': config.get('claude_model','claude-3-haiku-20240307'), 'max_tokens': 5, 'messages': [{'role':'user','content':'ping'}]}).encode(),
            headers={'Content-Type': 'application/json', 'x-api-key': api_key, 'anthropic-version': '2023-06-01'})
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
            return {'ok': True, 'model': data.get('model','?'), 'provider': 'claude'}
    except Exception as e:
        return {'ok': False, 'error': str(e)[:200]}


import io

def export_to_excel(df, sheet_name="Kalkulation", filename_prefix="export"):
    """Export a DataFrame to Excel bytes with formatting.
    
    Returns:
        BytesIO object ready for st.download_button()
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name max 31 chars
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
    
    # Auto-fit column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        max_width = max(
            len(str(col_name)),
            max((len(str(val)) for val in df[col_name].astype(str)), default=0)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 3, 40)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

